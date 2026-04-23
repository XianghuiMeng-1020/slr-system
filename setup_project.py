#!/usr/bin/env python3
"""
SLR System Data Setup Script
自动将编码方案和56篇文献导入系统，并创建三个编码者账号。

用法:
  python3 setup_project.py [--api-url http://localhost:8000]
"""
import sys
import json
import time
import argparse
import os
import openpyxl
import requests

# Bypass system proxy for localhost
os.environ.setdefault("no_proxy", "localhost,127.0.0.1")
_SESSION = requests.Session()
_SESSION.trust_env = False  # ignore system proxy settings

# ─── Configuration ─────────────────────────────────────────────────────────────

CODERS = [
    {"email": "coder1@slr.local", "password": "Slr2026#1", "label": "First Coder"},
    {"email": "coder2@slr.local", "password": "Slr2026#2", "label": "Second Coder (You)"},
    {"email": "coder3@slr.local", "password": "Slr2026#3", "label": "Third Coder"},
]

XLSX_PATH = "review_690366_extracted_data_xlsx_20260423131209.xlsx"
SCHEME_CSV_PATH = "coding_scheme_CSCL.csv"

# ─── Helpers ───────────────────────────────────────────────────────────────────

def ok(resp, label=""):
    if resp.status_code not in (200, 201):
        print(f"  ✗ {label}: HTTP {resp.status_code} — {resp.text[:300]}")
        return None
    data = resp.json().get("data")
    return data


def register_or_login(api, email, password):
    """Register; if already registered, login."""
    r = _SESSION.post(f"{api}/api/auth/register",
                      json={"email": email, "password": password})
    if r.status_code == 200:
        return r.json()["data"]["token"]
    # Already exists → login
    r = _SESSION.post(f"{api}/api/auth/login",
                      json={"email": email, "password": password})
    if r.status_code != 200:
        raise RuntimeError(f"Cannot login {email}: {r.text[:200]}")
    return r.json()["data"]["token"]


# ─── XLSX parsing ──────────────────────────────────────────────────────────────

SKIP_ROWS = {
    "Agreed", "Study Identification", "Methods", "Population",
    "Interventions", "Outcomes", "Linked suggestion", "Notes",
    "GenAI agent condition", "Control condition", "Overall",
    "", None,
}

def parse_studies(xlsx_path):
    """Return list of {short_name, title} from the Studies sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Studies"]
    studies = []
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1]:
            studies.append({
                "short_name": str(row[0]).strip(),
                "title": str(row[1]).strip(),
            })
    wb.close()
    return studies


def parse_study_coding(ws_name, wb):
    """
    Extract the 'Agreed' coding section from a study sheet.
    Returns a dict of {field: value}.
    """
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))

    # The agreed section is the first content block (before a coder name appears)
    KNOWN_CODER_NAMES = {"Yujing Zhang", "XIANGHUI MENG", "Yujing", "Meng"}
    coding = {}
    in_agreed = True

    for row in rows:
        field = row[0]
        if field is None or str(field).strip() in SKIP_ROWS:
            continue
        field_str = str(field).strip()
        # Stop parsing if we hit a coder's name row (individual coding begins)
        if field_str in KNOWN_CODER_NAMES:
            break
        val = row[1]
        val_str = str(val).strip() if val is not None else ""
        if val_str and val_str not in ("None", "\\"):
            coding[field_str] = val_str

    return coding


def build_document_text(short_name, title, coding):
    """Build a readable text summary to store as the 'document'."""
    lines = [
        f"# {title}",
        f"**Study ID:** {short_name}",
        "",
    ]
    sections = {
        "Study Identification": ["Country", "Publication type", "Journal sources", "Publication year"],
        "Methods": [
            "Study design", "Collaborative task",
            "Group collaboration environment/platform",
            "GenAI System & Model Architecture",
            "GenAI Agent's Instructional & Learning Design Theory",
            "Group interaction modality (Temporal modality+Communication）",
            "Interaction Data Sources", "Total Study Duration", "Task Duration",
            "Control condition",
            "Interaction Level-System Embedded analytic approaches",
            "Interaction Level-Post-hoc researcher analytic approaches",
            "Role of GenAI", "GenAI Integration Mode",
        ],
        "Population": [
            "Group composition (number of human agents and GenAI agents)",
            "Participant's Course / discipline (e.g., engineering, teacher education)",
        ],
    }
    for section, fields in sections.items():
        lines.append(f"\n## {section}")
        for f in fields:
            v = coding.get(f, "")
            if v:
                lines.append(f"**{f}:** {v}")

    # Any remaining keys not in sections
    covered = {f for fs in sections.values() for f in fs}
    extras = {k: v for k, v in coding.items() if k not in covered}
    if extras:
        lines.append("\n## Other")
        for k, v in extras.items():
            lines.append(f"**{k}:** {v}")

    return "\n".join(lines)


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--api-url", default="http://localhost:8000")
    args = parser.parse_args()
    api = args.api_url.rstrip("/")

    print(f"\n{'='*60}")
    print(f"SLR System Setup — {api}")
    print(f"{'='*60}\n")

    # ── 1. Health check ────────────────────────────────────────────
    print("① Checking backend health…")
    try:
        r = _SESSION.get(f"{api}/api/health", timeout=5)
        h = r.json().get("data", {})
        print(f"   DB: {h.get('db')}  |  disk_free: {h.get('disk_free_mb')} MB")
    except Exception as e:
        print(f"   ✗ Cannot reach {api}: {e}")
        print("   → Make sure the backend is running (cd backend && uvicorn main:app --reload)")
        sys.exit(1)

    # ── 2. Register / login all coders ────────────────────────────
    print("\n② Creating coder accounts…")
    tokens = {}
    for coder in CODERS:
        try:
            tok = register_or_login(api, coder["email"], coder["password"])
            tokens[coder["email"]] = tok
            print(f"   ✓ {coder['label']}: {coder['email']} / {coder['password']}")
        except Exception as e:
            print(f"   ✗ {coder['email']}: {e}")

    # Use second coder as project owner
    owner_email = CODERS[1]["email"]
    owner_token = tokens[owner_email]
    headers = {"Authorization": f"Bearer {owner_token}"}

    # ── 3. Create project ─────────────────────────────────────────
    print("\n③ Creating project…")
    r = _SESSION.post(f"{api}/api/projects",
                      json={"mode": "theme-verification"},
                      headers=headers)
    proj = ok(r, "create project")
    if not proj:
        sys.exit(1)
    project_id = proj["id"]
    print(f"   ✓ Project ID: {project_id}")

    # Save project ID for reference
    with open("project_id.txt", "w") as f:
        f.write(project_id)

    # ── 4. Enable dual coding ─────────────────────────────────────
    print("\n④ Enabling dual-blind coding mode…")
    r = _SESSION.put(f"{api}/api/projects/{project_id}/settings",
                     json={"dual_coding_blind": True},
                     headers=headers)
    ok(r, "settings")
    print("   ✓ Dual-blind coding enabled")

    # ── 5. Upload coding scheme ───────────────────────────────────
    print("\n⑤ Uploading coding scheme from CSV…")
    with open(SCHEME_CSV_PATH, "rb") as f:
        r = _SESSION.post(
            f"{api}/api/projects/{project_id}/coding-scheme",
            files={"file": ("coding_scheme_CSCL.csv", f, "text/csv")},
            headers=headers,
        )
    scheme = ok(r, "coding scheme")
    if scheme:
        print(f"   ✓ {len(scheme)} coding scheme items uploaded")
    else:
        print("   ✗ Failed to upload coding scheme")
        sys.exit(1)

    # ── 6. Import studies ─────────────────────────────────────────
    print("\n⑥ Importing 56 studies from XLSX…")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    studies = parse_studies(XLSX_PATH)
    wb_full = openpyxl.load_workbook(XLSX_PATH, read_only=True)

    imported = 0
    failed = 0
    doc_ids = {}

    for study in studies:
        short = study["short_name"]
        title = study["title"]

        # Try to get existing coding from the sheet
        coding = {}
        if short in wb_full.sheetnames:
            try:
                coding = parse_study_coding(short, wb_full)
            except Exception:
                pass

        text_content = build_document_text(short, title, coding)
        fname = f"{short}.txt"

        r = _SESSION.post(
            f"{api}/api/projects/{project_id}/import-text",
            files={"file": (fname, text_content.encode("utf-8"), "text/plain")},
            headers=headers,
        )
        result = ok(r, f"import {short}")
        if result:
            doc_ids[short] = result["id"]
            imported += 1
            if imported % 10 == 0:
                print(f"   … {imported}/{len(studies)} imported")
        else:
            failed += 1

    wb_full.close()
    print(f"   ✓ {imported} studies imported  ({failed} failed)")

    # Save doc mapping
    with open("doc_ids.json", "w") as f:
        json.dump(doc_ids, f, ensure_ascii=False, indent=2)

    # ── 7. Add other coders to project ───────────────────────────
    print("\n⑦ Adding coders as project members…")
    for coder in CODERS:
        if coder["email"] == owner_email:
            continue
        r = _SESSION.post(
            f"{api}/api/projects/{project_id}/members",
            params={"email": coder["email"]},
            headers=headers,
        )
        result = ok(r, f"add {coder['label']}")
        if result:
            print(f"   ✓ Added: {coder['label']} ({coder['email']})")
        else:
            print(f"   ✗ Could not add {coder['label']}")

    # ── 8. Summary ────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print("✅ SETUP COMPLETE")
    print(f"{'='*60}")
    print(f"\n  Project ID : {project_id}")
    print(f"  Studies    : {imported} imported")
    print(f"  Coding items: {len(scheme) if scheme else '?'}")
    print(f"\n  ── Coder Login Credentials ──")
    for coder in CODERS:
        print(f"  {coder['label']:25s}  {coder['email']}  /  {coder['password']}")
    print(f"\n  Share the URL + credentials with your team.")
    print(f"  Project ID saved to: project_id.txt")
    print()


if __name__ == "__main__":
    main()
